import os
from io import BytesIO

from openpyxl import load_workbook, Workbook

from app.services.file_services.base_service import serialize_result, deserialize_result
from app.services.file_services.csv_service import CSVFileService
from app.utils.cache import cache
from app.utils.md5 import calculate_md5_from_data_and_pattern


class XLSXFileService(CSVFileService):
    def __init__(self, storage_path, search_word):
        super().__init__(storage_path, search_word)


    def _save_rows_to_file(self, headers, rows, file_name):
        file_path = os.path.join(self.storage_path, f"{file_name}.xlsx")
        work_book = Workbook()
        work_sheet = work_book.active
        work_sheet.append(headers)
        for row in rows:
            work_sheet.append(row)
        work_book.save(file_path)
        return f'file://{os.path.abspath(file_path)}'


    @cache(key_func=calculate_md5_from_data_and_pattern, serializer=serialize_result, deserializer=deserialize_result)
    def _process_file(self, data, pattern):
        file_in_memory = BytesIO(data)
        work_book = load_workbook(file_in_memory)
        work_sheet = work_book.active

        headers = [cell.value for cell in work_sheet[1]]
        rows = list(work_sheet.iter_rows(min_row=2, values_only=True))

        valid_column_index = self._get_column_index(headers, 'Company Name')
        is_word_found, valid_rows, invalid_rows = self._process_rows(rows, valid_column_index, pattern)

        file_uuid = uuid.uuid4()
        output_valid_file = self._save_rows_to_file(headers, valid_rows, f'valid_{file_uuid}')
        output_invalid_file = self._save_rows_to_file(headers, invalid_rows, f'invalid_{file_uuid}')

        return is_word_found, output_valid_file, output_invalid_file
