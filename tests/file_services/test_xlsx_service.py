import os
import shutil

import pytest
from openpyxl.reader.excel import load_workbook

from app.services.file_services.xlsx_service import XLSXFileService


@pytest.fixture(scope="module")
def temp_dir():
    dir_path = "tests/tmp"
    os.makedirs(dir_path, exist_ok=True)
    yield dir_path

    shutil.rmtree(dir_path)


@pytest.mark.parametrize("file_path, pattern, expected_valid, expected_invalid", [
    ("example.xlsx", "Quantori", "example_valid.xlsx", "example_invalid.xlsx"),
    ("example.xlsx", "quantori", "example_valid.xlsx", "example_invalid.xlsx"),
    ("example.xlsx", "NotFound", "not_found_valid.xlsx", "not_found_invalid.xlsx"),
    ("example.xlsx", "", "all.xlsx", "empty_data.xlsx"),
    ("empty_data.xlsx", "quantori", "empty_data.xlsx", "empty_data.xlsx"),
])
def test_process_file(temp_dir, file_path, pattern, expected_valid, expected_invalid):
    service = XLSXFileService(temp_dir, pattern)

    with open(f'tests/data/{file_path}', 'rb') as f:
        _, result_valid_file, result_invalid_file = service._process_file(f.read(), service.search_word)

    assert os.path.exists(result_valid_file[(len('file://')):])
    assert os.path.exists(result_invalid_file[(len('file://')):])

    work_book_valid_expected = load_workbook(f'tests/data/{expected_valid}')
    work_sheet_valid_expected = work_book_valid_expected.active
    work_book_invalid_expected = load_workbook(f'tests/data/{expected_invalid}')
    work_sheet_invalid_expected = work_book_invalid_expected.active
    work_book_valid = load_workbook(result_valid_file[(len('file://')):])
    work_sheet_valid = work_book_valid.active
    work_book_invalid = load_workbook(result_invalid_file[(len('file://')):])
    work_sheet_invalid = work_book_invalid.active

    compare_sheets_cell_by_cell(work_sheet_valid, work_sheet_valid_expected)
    compare_sheets_cell_by_cell(work_sheet_invalid, work_sheet_invalid_expected)


def compare_sheets_cell_by_cell(sheet1, sheet2):
    assert sheet1.max_row == sheet2.max_row
    assert sheet1.max_column == sheet2.max_column

    for row in range(1, sheet1.max_row + 1):
        for col in range(1, sheet1.max_column + 1):
            cell1 = sheet1.cell(row=row, column=col).value
            cell2 = sheet2.cell(row=row, column=col).value
            assert cell1 == cell2